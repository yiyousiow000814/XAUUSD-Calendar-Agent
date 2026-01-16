import argparse
import collections
import os
import random
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# When executing this file via `python scripts/calendar/economic_calendar_fetcher.py`,
# Python sets sys.path[0] to `scripts/calendar`, so `import scripts...` fails unless
# we add the repository root explicitly.
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from scripts.calendar import calendar_processing as processing  # noqa: E402
from scripts.calendar import calendar_pruning  # noqa: E402

# Base folder to store generated artifacts under the repository's data directory
BASE_DATA_DIR = REPO_ROOT / "data"
_DEFAULT_CALENDAR_DIR = BASE_DATA_DIR / "Economic_Calendar"
CALENDAR_OUTPUT_DIR = Path(os.getenv("CALENDAR_OUTPUT_DIR", str(_DEFAULT_CALENDAR_DIR)))
if not CALENDAR_OUTPUT_DIR.is_absolute():
    CALENDAR_OUTPUT_DIR = REPO_ROOT / CALENDAR_OUTPUT_DIR
YEARLY_OUTPUT_DIR = CALENDAR_OUTPUT_DIR

_URL_TOKEN_RE = re.compile(r"(?:https?://|www\\.)\\S+", re.IGNORECASE)
_DOMAIN_TOKEN_RE = re.compile(
    r"\\b(?:[A-Za-z0-9-]{2,63}\\.)+[A-Za-z]{2,24}\\b", re.IGNORECASE
)


def _sanitize_text_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = _URL_TOKEN_RE.sub("", value)
    cleaned = _DOMAIN_TOKEN_RE.sub("", cleaned)
    cleaned = re.sub(r"\\s{2,}", " ", cleaned).strip()
    return cleaned


def _load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip()
        if (value.startswith('"') and value.endswith('"')) or (
            value.startswith("'") and value.endswith("'")
        ):
            value = value[1:-1]
        if not key:
            continue
        if os.environ.get(key, "").strip():
            continue
        os.environ[key] = value


_load_dotenv(REPO_ROOT / ".env")
_load_dotenv(REPO_ROOT / "user-data" / ".env")


def _load_env_from_text_file(var_name: str, default_path: Path) -> None:
    if os.environ.get(var_name, "").strip():
        return

    file_path = os.getenv(f"{var_name}_FILE")
    path = Path(file_path).expanduser() if file_path else default_path
    if not path.exists():
        return

    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        value = raw_line.lstrip("\ufeff").strip()
        if not value or value.startswith("#"):
            continue
        if value.startswith("-"):
            value = value[1:].strip()
        if value:
            os.environ[var_name] = value
            return


_load_env_from_text_file(
    "CALENDAR_API_ENDPOINT", REPO_ROOT / "user-data" / "CALENDAR_API_ENDPOINT.txt"
)
_load_env_from_text_file(
    "CALENDAR_REFERER", REPO_ROOT / "user-data" / "CALENDAR_REFERER.txt"
)

ECON_CALENDAR_ENDPOINT = os.getenv("CALENDAR_API_ENDPOINT", "").strip()
if not ECON_CALENDAR_ENDPOINT:
    raise SystemExit(
        "Missing CALENDAR_API_ENDPOINT. Provide it via environment variable or "
        "user-data/CALENDAR_API_ENDPOINT.txt."
    )

CALENDAR_REFERER = os.getenv("CALENDAR_REFERER", "").strip()
if not CALENDAR_REFERER:
    raise SystemExit(
        "Missing CALENDAR_REFERER. Provide it via environment variable or "
        "user-data/CALENDAR_REFERER.txt."
    )
ECON_CALENDAR_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": CALENDAR_REFERER,
}
DEFAULT_IMPORTANCE = (1, 2, 3)
# Calendar provider timezone id for GMT+08:00 (Singapore/Kuala Lumpur). The value
# can be overridden via the ``CALENDAR_TIMEZONE_ID`` environment variable.
_timezone_raw = (os.getenv("CALENDAR_TIMEZONE_ID") or "").strip()
try:
    DEFAULT_TIMEZONE_ID = int(_timezone_raw or "178")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_TIMEZONE_ID={_timezone_raw!r}; falling back to 178."
    )
    DEFAULT_TIMEZONE_ID = 178
ECON_CALENDAR_PAGE_SIZE = 200
_page_delay_min_raw = (os.getenv("CALENDAR_PAGE_DELAY_MIN_SECONDS") or "").strip()
_page_delay_max_raw = (os.getenv("CALENDAR_PAGE_DELAY_MAX_SECONDS") or "").strip()
_page_delay_raw = (os.getenv("CALENDAR_PAGE_DELAY_SECONDS") or "").strip()

PAGE_DELAY_MIN_SECONDS = 5.0
PAGE_DELAY_MAX_SECONDS = 7.0
if _page_delay_min_raw or _page_delay_max_raw:
    try:
        PAGE_DELAY_MIN_SECONDS = float(_page_delay_min_raw or "5")
        PAGE_DELAY_MAX_SECONDS = float(_page_delay_max_raw or "7")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_PAGE_DELAY_MIN_SECONDS/CALENDAR_PAGE_DELAY_MAX_SECONDS="
            f"{_page_delay_min_raw!r}/{_page_delay_max_raw!r}; falling back to 5..7."
        )
        PAGE_DELAY_MIN_SECONDS = 5.0
        PAGE_DELAY_MAX_SECONDS = 7.0
elif _page_delay_raw:
    # Backward compatible fixed delay.
    try:
        PAGE_DELAY_MIN_SECONDS = PAGE_DELAY_MAX_SECONDS = float(_page_delay_raw)
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_PAGE_DELAY_SECONDS={_page_delay_raw!r}; "
            "falling back to 5..7."
        )
        PAGE_DELAY_MIN_SECONDS = 5.0
        PAGE_DELAY_MAX_SECONDS = 7.0

PAGE_DELAY_MIN_SECONDS = max(0.0, min(PAGE_DELAY_MIN_SECONDS, PAGE_DELAY_MAX_SECONDS))
PAGE_DELAY_MAX_SECONDS = max(PAGE_DELAY_MIN_SECONDS, PAGE_DELAY_MAX_SECONDS)

_guard_ratio_raw = (os.getenv("CALENDAR_PRUNE_GUARD_RATIO") or "").strip()
try:
    PRUNE_GUARD_RATIO = float(_guard_ratio_raw or "0.6")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_PRUNE_GUARD_RATIO={_guard_ratio_raw!r}; "
        "falling back to 0.6."
    )
    PRUNE_GUARD_RATIO = 0.6

_guard_min_raw = (os.getenv("CALENDAR_PRUNE_GUARD_MIN_NEW_NONHOLIDAY") or "").strip()
try:
    PRUNE_GUARD_MIN_NEW_NONHOLIDAY = int(_guard_min_raw or "5")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_PRUNE_GUARD_MIN_NEW_NONHOLIDAY={_guard_min_raw!r}; "
        "falling back to 5."
    )
    PRUNE_GUARD_MIN_NEW_NONHOLIDAY = 5

_min_interval_raw = (os.getenv("CALENDAR_HTTP_MIN_INTERVAL_SECONDS") or "").strip()
try:
    HTTP_MIN_INTERVAL_SECONDS = float(_min_interval_raw or "0")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_MIN_INTERVAL_SECONDS={_min_interval_raw!r}; "
        "falling back to 0."
    )
    HTTP_MIN_INTERVAL_SECONDS = 0.0

HTTP_STATS_ENABLED = (os.getenv("CALENDAR_HTTP_STATS") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
_REQUEST_TIMES = collections.deque(maxlen=5000)  # monotonic seconds
_LAST_REQUEST_AT: float | None = None
_http_attempts_raw = (os.getenv("CALENDAR_HTTP_MAX_ATTEMPTS") or "").strip()
try:
    HTTP_MAX_ATTEMPTS = int(_http_attempts_raw or "3")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_MAX_ATTEMPTS={_http_attempts_raw!r}; "
        "falling back to 3."
    )
    HTTP_MAX_ATTEMPTS = 3

_cooldown_raw = (os.getenv("CALENDAR_HTTP_RATE_LIMIT_COOLDOWN_SECONDS") or "").strip()
try:
    HTTP_RATE_LIMIT_COOLDOWN_SECONDS = int(_cooldown_raw or "320")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_RATE_LIMIT_COOLDOWN_SECONDS={_cooldown_raw!r}; "
        "falling back to 320."
    )
    HTTP_RATE_LIMIT_COOLDOWN_SECONDS = 320

_cooldowns_raw = (os.getenv("CALENDAR_HTTP_RATE_LIMIT_MAX_COOLDOWNS") or "").strip()
try:
    HTTP_RATE_LIMIT_MAX_COOLDOWNS = int(_cooldowns_raw or "0")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_RATE_LIMIT_MAX_COOLDOWNS={_cooldowns_raw!r}; "
        "falling back to 0."
    )
    HTTP_RATE_LIMIT_MAX_COOLDOWNS = 0

_day_delay_raw = (os.getenv("CALENDAR_DAY_DELAY_SECONDS") or "").strip() or (
    os.getenv("CALENDAR_DAY_DELAY") or ""
).strip()
DAY_DELAY_SECONDS = None
if _day_delay_raw:
    try:
        DAY_DELAY_SECONDS = float(_day_delay_raw)
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_DAY_DELAY_SECONDS={_day_delay_raw!r}; "
            "disabling fixed day delay."
        )

_day_delay_min_raw = (os.getenv("CALENDAR_DAY_DELAY_MIN_SECONDS") or "").strip()
try:
    DAY_DELAY_MIN_SECONDS = float(_day_delay_min_raw or "1")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_DAY_DELAY_MIN_SECONDS={_day_delay_min_raw!r}; "
        "falling back to 1."
    )
    DAY_DELAY_MIN_SECONDS = 1.0

_day_delay_max_raw = (os.getenv("CALENDAR_DAY_DELAY_MAX_SECONDS") or "").strip()
try:
    DAY_DELAY_MAX_SECONDS = float(_day_delay_max_raw or "3")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_DAY_DELAY_MAX_SECONDS={_day_delay_max_raw!r}; "
        "falling back to 3."
    )
    DAY_DELAY_MAX_SECONDS = 3.0


_jitter_min_raw = (os.getenv("CALENDAR_HTTP_JITTER_MIN_SECONDS") or "").strip()
try:
    HTTP_JITTER_MIN_SECONDS = float(_jitter_min_raw or "0.8")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_JITTER_MIN_SECONDS={_jitter_min_raw!r}; "
        "falling back to 0.8."
    )
    HTTP_JITTER_MIN_SECONDS = 0.8

_jitter_max_raw = (os.getenv("CALENDAR_HTTP_JITTER_MAX_SECONDS") or "").strip()
try:
    HTTP_JITTER_MAX_SECONDS = float(_jitter_max_raw or "1.5")
except ValueError:
    print(
        f"[WARNING] Invalid CALENDAR_HTTP_JITTER_MAX_SECONDS={_jitter_max_raw!r}; "
        "falling back to 1.5."
    )
    HTTP_JITTER_MAX_SECONDS = 1.5


def _sleep_request_jitter() -> None:
    if HTTP_JITTER_MAX_SECONDS <= 0:
        return
    low = max(0.0, min(HTTP_JITTER_MIN_SECONDS, HTTP_JITTER_MAX_SECONDS))
    high = max(0.0, max(HTTP_JITTER_MIN_SECONDS, HTTP_JITTER_MAX_SECONDS))
    if high <= 0:
        return
    time.sleep(random.uniform(low, high))


def _sleep_min_interval() -> None:
    if HTTP_MIN_INTERVAL_SECONDS <= 0:
        return
    now = time.monotonic()
    if _LAST_REQUEST_AT is None:
        return
    elapsed = now - _LAST_REQUEST_AT
    remaining = HTTP_MIN_INTERVAL_SECONDS - elapsed
    if remaining > 0:
        time.sleep(remaining)


def _record_request() -> None:
    global _LAST_REQUEST_AT
    now = time.monotonic()
    _LAST_REQUEST_AT = now
    _REQUEST_TIMES.append(now)


def _requests_in_last(seconds: int) -> int:
    if seconds <= 0:
        return 0
    cutoff = time.monotonic() - float(seconds)
    # Deque is ordered; count from left until cutoff.
    count = 0
    for ts in reversed(_REQUEST_TIMES):
        if ts < cutoff:
            break
        count += 1
    return count


def _log_http_stats(prefix: str, *, status: int | None = None) -> None:
    if not HTTP_STATS_ENABLED and status != 429:
        return
    last_60 = _requests_in_last(60)
    last_300 = _requests_in_last(300)
    extra = f" status={status}" if status is not None else ""
    print(f"[HTTP] {prefix}: last_60s={last_60} last_5m={last_300}{extra}")


def _sleep_day_delay() -> None:
    if DAY_DELAY_SECONDS is not None:
        if DAY_DELAY_SECONDS > 0:
            time.sleep(DAY_DELAY_SECONDS)
        return

    low = max(0.0, min(DAY_DELAY_MIN_SECONDS, DAY_DELAY_MAX_SECONDS))
    high = max(0.0, max(DAY_DELAY_MIN_SECONDS, DAY_DELAY_MAX_SECONDS))
    if high <= 0:
        return
    time.sleep(random.uniform(low, high))


class CalendarFetchError(RuntimeError):
    def __init__(
        self,
        message: str,
        failed_start: datetime,
        failed_end: datetime,
        partial_headers: list[str],
        partial_rows: list[list[str]],
    ) -> None:
        super().__init__(message)
        self.failed_start = failed_start
        self.failed_end = failed_end
        self.partial_headers = partial_headers
        self.partial_rows = partial_rows


def enforce_backup_limit(directory, limit=15):
    """Keep only the most recent `limit` backup files in `directory`."""
    try:
        backups = [
            os.path.join(directory, name)
            for name in os.listdir(directory)
            if os.path.isfile(os.path.join(directory, name))
        ]
        if len(backups) < limit:
            return

        backups.sort(key=os.path.getmtime)
        excess = len(backups) - limit + 1
        for obsolete_path in backups[:excess]:
            try:
                os.remove(obsolete_path)
                print(f"[INFO] Removed obsolete backup: {obsolete_path}")
            except OSError as exc:
                print(f"[WARNING] Failed to delete backup {obsolete_path}: {exc}")
    except FileNotFoundError:
        # Directory may not exist yet; nothing to prune.
        return


def export_yearly_breakdown(df, website_prefix, file_name, changed_years=None):
    """Write per-year snapshots under the data directory.

    When ``changed_years`` is provided, only the specified years are refreshed to
    avoid touching historical files unnecessarily.
    """
    if df.empty or "Date" not in df.columns:
        return

    YEARLY_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Parse the Date column to datetime for reliable grouping
    working_df = df.copy()
    working_df["Date"] = pd.to_datetime(working_df["Date"], errors="coerce")
    working_df = working_df.dropna(subset=["Date"])

    if working_df.empty:
        return

    available_years = {
        int(year) for year in working_df["Date"].dt.year.dropna().unique()
    }
    if changed_years is not None:
        normalized_years = {int(year) for year in changed_years}
        target_years = sorted(available_years & normalized_years)
        if not target_years:
            return
    else:
        target_years = sorted(available_years)

    target_set = set(target_years)

    for year, group in working_df.groupby(working_df["Date"].dt.year):
        year_int = int(year)
        if year_int not in target_set:
            continue
        year_dir = YEARLY_OUTPUT_DIR / str(year_int)
        year_dir.mkdir(parents=True, exist_ok=True)

        export_df = group.sort_values("Date").copy()
        export_df["Date"] = export_df["Date"].dt.strftime("%Y-%m-%d")
        export_df = processing.sort_calendar_dataframe(export_df)

        excel_path = year_dir / f"{year_int}_calendar.xlsx"
        csv_path = year_dir / f"{year_int}_calendar.csv"
        json_path = year_dir / f"{year_int}_calendar.json"

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Data")

        export_df.to_csv(csv_path, index=False)
        export_df.to_json(json_path, orient="records", indent=4)

        print(
            f"[INFO] Yearly export generated for {year_int}: {excel_path}, {csv_path}, {json_path}"
        )


def day_range(start_date: datetime, end_date: datetime):
    """Yield (start, end) datetimes for each day between the boundaries."""
    current = start_date
    one_day = timedelta(days=1)
    while current <= end_date:
        yield current, current
        current += one_day


def chunk_date_range(start_date: datetime, end_date: datetime, chunk_days: int):
    """Yield inclusive (start, end) windows covering the date range."""
    if chunk_days <= 0:
        raise ValueError("chunk_days must be a positive integer.")

    current = start_date
    while current <= end_date:
        chunk_end = min(end_date, current + timedelta(days=chunk_days - 1))
        yield current, chunk_end
        current = chunk_end + timedelta(days=1)


def sort_calendar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Sort ``df`` by date and the semantic meaning of the ``Time`` column."""

    if df.empty or "Date" not in df.columns or "Time" not in df.columns:
        return df

    working = df.copy()

    working["_sort_date"] = pd.to_datetime(working["Date"], errors="coerce")
    time_str = working["Time"].fillna("").astype(str).str.strip()

    working["_is_all_day"] = time_str.str.lower() == "all day"
    working["_time_parsed"] = pd.to_datetime(time_str, format="%H:%M", errors="coerce")

    working["_sort_bucket"] = 2  # textual placeholders (e.g. Tentative/TBA)
    working.loc[working["_is_all_day"], "_sort_bucket"] = 0
    working.loc[working["_time_parsed"].notna(), "_sort_bucket"] = 1
    working.loc[
        time_str.eq("") | time_str.str.lower().isin(["nan", "none"]), "_sort_bucket"
    ] = 3

    tie_breakers = [
        column
        for column in ["Event", "Cur.", "Imp.", "Actual", "Forecast", "Previous"]
        if column in working.columns
    ]

    sort_columns = ["_sort_date", "_sort_bucket", "_time_parsed", "Time", *tie_breakers]
    ascending_flags = [True] * len(sort_columns)

    working.sort_values(
        by=sort_columns,
        ascending=ascending_flags,
        inplace=True,
        kind="stable",
    )

    working.drop(
        columns=["_sort_date", "_sort_bucket", "_time_parsed", "_is_all_day"],
        inplace=True,
        errors="ignore",
    )

    return working


COLUMN_WIDTH_OVERRIDES = {
    "Date": 9.71,
    "Day": 10.71,
    "Time": 4.86,
    "Cur.": 4.57,
    "Imp.": 7.71,
    "Event": 50.0,
    "Actual": 8.57,
    "Forecast": 8.57,
    "Previous": 8.57,
}

KEY_COLUMNS = ["Date", "Time", "Cur.", "Event"]


def _parse_time_minutes(value: object) -> float | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    try:
        parsed = datetime.strptime(text, "%H:%M")
    except ValueError:
        return None
    return float(parsed.hour * 60 + parsed.minute)


def _format_time_minutes(minutes: float) -> str:
    total = int(round(minutes))
    total = max(0, min(total, 23 * 60 + 59))
    hour = total // 60
    minute = total % 60
    return f"{hour:02d}:{minute:02d}"


def _is_missing_token(value: object, *, missing_tokens: set[str]) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    text = str(value).strip()
    if not text:
        return True
    return text.lower() in missing_tokens


def _choose_canonical_time(times: list[float]) -> float | None:
    if not times:
        return None

    def score(t: float) -> tuple[int, int, float]:
        minute = int(round(t)) % 60
        is_hour_or_half = 1 if minute in {0, 30} else 0
        is_quarter = 1 if minute in {0, 15, 30, 45} else 0
        return (is_hour_or_half, is_quarter, t)

    return max(times, key=score)


def _snap_time_to_canonical_minutes(minutes: float, *, threshold_minutes: int) -> float:
    if threshold_minutes <= 0:
        return minutes
    if pd.isna(minutes):
        return minutes
    rounded = int(round(minutes))
    if rounded < 0 or rounded > 23 * 60 + 59:
        return minutes

    canonical = list(range(0, 24 * 60, 5))
    nearest = min(canonical, key=lambda m: abs(m - rounded))
    if abs(nearest - rounded) <= threshold_minutes:
        return float(nearest)
    return minutes


def _apply_fuzzy_time_dedup(
    df: pd.DataFrame,
    *,
    group_columns: list[str],
    threshold_minutes: int,
) -> pd.DataFrame:
    if df.empty or threshold_minutes <= 0:
        return df

    if any(col not in df.columns for col in group_columns):
        return df

    working = df.copy()
    working["_time_minutes"] = working.get("Time", "").map(_parse_time_minutes)

    missing_tokens = {"tba", "tentative", "n/a", "na"}

    kept_rows: list[pd.Series] = []
    for _, group in working.groupby(group_columns, dropna=False, sort=False):
        if len(group) == 1:
            row = group.iloc[0].copy()
            row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(row)
            continue

        timed = group[group["_time_minutes"].notna()].copy()
        untimed = group[group["_time_minutes"].isna()].copy()

        if timed.empty:
            for _, row in group.iterrows():
                row = row.copy()
                row.drop(labels=["_time_minutes"], inplace=True)
                kept_rows.append(row)
            continue

        timed = timed.sort_values(by="_time_minutes", kind="mergesort")
        clusters: list[list[int]] = []
        current: list[int] = []
        last_time: float | None = None
        for idx, row in timed.iterrows():
            t = row["_time_minutes"]
            if not isinstance(t, float):
                continue
            if last_time is None or abs(t - last_time) <= threshold_minutes:
                current.append(int(idx))
            else:
                clusters.append(current)
                current = [int(idx)]
            last_time = t
        if current:
            clusters.append(current)

        for cluster in clusters:
            cluster_df = timed.loc[cluster].copy()
            # choose the row with highest completeness_score
            best_idx = int(
                cluster_df.sort_values(
                    by=["completeness_score"],
                    ascending=False,
                    kind="mergesort",
                ).index[0]
            )
            best_row = cluster_df.loc[best_idx].copy()

            canonical_time = _choose_canonical_time(
                [
                    float(x)
                    for x in cluster_df["_time_minutes"].tolist()
                    if isinstance(x, float)
                ]
            )
            if canonical_time is not None:
                best_row["Time"] = _format_time_minutes(canonical_time)

            # fill missing values from other rows in the cluster
            for _, candidate in cluster_df.iterrows():
                if int(candidate.name) == best_idx:
                    continue
                for col_name in cluster_df.columns:
                    if col_name in {"_time_minutes"}:
                        continue
                    if _is_missing_token(
                        best_row.get(col_name), missing_tokens=missing_tokens
                    ) and not _is_missing_token(
                        candidate.get(col_name), missing_tokens=missing_tokens
                    ):
                        best_row[col_name] = candidate.get(col_name)

            best_row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(best_row)

        for _, row in untimed.iterrows():
            row = row.copy()
            row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(row)

    result = pd.DataFrame(kept_rows)
    return result.reset_index(drop=True)


def merge_calendar_frames(
    existing_df: pd.DataFrame, new_df: pd.DataFrame
) -> pd.DataFrame:
    working = pd.concat([existing_df, new_df], ignore_index=True, sort=False)
    working.replace(["nan", "NaN", "None"], pd.NA, inplace=True)
    if "Time" in working.columns:
        working["Time"] = working["Time"].fillna("").astype(str).str.strip()
    else:
        working["Time"] = ""

    for col_name in working.columns:
        if working[col_name].dtype == object:
            working[col_name] = working[col_name].map(_sanitize_text_value)

    working["Date_dt"] = pd.to_datetime(working.get("Date"), errors="coerce")
    working = working.dropna(subset=["Date_dt"])  # type: ignore[arg-type]

    completeness = pd.Series(0, index=working.index, dtype="int64")
    missing_tokens = {"tba", "tentative", "n/a", "na"}
    for col_name in working.columns:
        col = working[col_name]
        present = col.notna()
        if col.dtype == object:
            text = col.astype(str).str.strip()
            present &= text.ne("") & ~text.str.lower().isin(missing_tokens)
        completeness += present.astype("int64")

    working["completeness_score"] = completeness
    working = working.sort_values(by="completeness_score", ascending=False)
    working = working.drop_duplicates(subset=KEY_COLUMNS, keep="first")

    _fuzzy_raw = (os.getenv("CALENDAR_EVENT_TIME_FUZZY_DEDUP_MINUTES") or "").strip()
    try:
        fuzzy_minutes = int(_fuzzy_raw or "2")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_EVENT_TIME_FUZZY_DEDUP_MINUTES={_fuzzy_raw!r}; "
            "falling back to 2."
        )
        fuzzy_minutes = 2

    working = _apply_fuzzy_time_dedup(
        working,
        group_columns=["Date", "Cur.", "Event"],
        threshold_minutes=fuzzy_minutes,
    )
    working.drop(columns=["completeness_score"], inplace=True)

    _snap_raw = (os.getenv("CALENDAR_TIME_SNAP_THRESHOLD_MINUTES") or "").strip()
    try:
        snap_threshold = int(_snap_raw or "2")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_TIME_SNAP_THRESHOLD_MINUTES={_snap_raw!r}; "
            "falling back to 2."
        )
        snap_threshold = 2

    if snap_threshold > 0 and "Time" in working.columns:
        minutes = working["Time"].map(_parse_time_minutes)
        working["Time"] = [
            (
                _format_time_minutes(
                    _snap_time_to_canonical_minutes(m, threshold_minutes=snap_threshold)
                )
                if isinstance(m, float) and not pd.isna(m)
                else t
            )
            for t, m in zip(working["Time"].tolist(), minutes.tolist())
        ]

    working["Date"] = working["Date_dt"].dt.strftime("%Y-%m-%d")
    working["Day"] = working["Date_dt"].dt.strftime("%A")
    working.drop(columns=["Date_dt"], inplace=True)

    working = processing.sort_calendar_dataframe(working)
    return working.reset_index(drop=True)


def write_calendar_outputs(df: pd.DataFrame, excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        worksheet = writer.sheets["Data"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_text_length = df[col_name].astype(str).map(len).max() or 0
            override_width = COLUMN_WIDTH_OVERRIDES.get(col_name)
            col_width = (
                override_width
                if override_width is not None
                else max(max_text_length + 2, 15)
            )

            cell_alignment = (
                Alignment(horizontal="left")
                if col_name == "Event"
                else Alignment(horizontal="center")
            )

            worksheet.column_dimensions[column_letter].width = col_width
            for cell in worksheet[column_letter]:
                cell.alignment = cell_alignment

        worksheet.freeze_panes = "A2"

        highlight_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        previous_date = None
        for row_idx, row in enumerate(df.itertuples(), start=2):
            current_row_date = row.Date
            if current_row_date != previous_date:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = highlight_fill
                previous_date = current_row_date

    df.to_csv(excel_path.with_suffix(".csv"), index=False)
    df.to_json(excel_path.with_suffix(".json"), orient="records", indent=4)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Fetch economic calendar data via HTTP API."
    )
    parser.add_argument(
        "--start-date",
        type=str,
        help="Start date in YYYY-MM-DD. Overrides --start-year when provided.",
    )
    parser.add_argument(
        "--end-date",
        type=str,
        help="End date in YYYY-MM-DD. Overrides --end-year when provided.",
    )
    parser.add_argument(
        "--start-year",
        type=int,
        help="Start year (defaults to current year if not provided).",
    )
    parser.add_argument(
        "--end-year",
        type=int,
        help="End year (defaults to start year when omitted).",
    )
    parser.add_argument(
        "--prune-existing-in-range",
        dest="prune_existing_in_range",
        action="store_true",
        help=(
            "Treat the fetched date window as authoritative by pruning existing rows "
            "inside the window before merging. A per-day guard prevents accidental "
            "data loss when upstream results are incomplete."
        ),
    )
    parser.add_argument(
        "--no-prune-existing-in-range",
        dest="prune_existing_in_range",
        action="store_false",
        help="Disable pruning existing rows inside the fetched date window.",
    )
    parser.set_defaults(prune_existing_in_range=None)
    return parser.parse_args()


def resolve_date_range(args):
    if args.start_date or args.end_date:
        if not (args.start_date and args.end_date):
            raise ValueError(
                "Both --start-date and --end-date must be provided together."
            )
        start_date = datetime.strptime(args.start_date, "%Y-%m-%d")
        end_date = datetime.strptime(args.end_date, "%Y-%m-%d")
    else:
        current_year = datetime.now().year
        start_year = args.start_year or current_year
        end_year = args.end_year or start_year
        start_date = datetime(start_year, 1, 1)
        end_date = datetime(end_year, 12, 31)

    if end_date < start_date:
        raise ValueError("End date must be on or after start date.")

    return start_date, end_date


def parse_calendar_html(html_snippet):
    """Parse calendar HTML snippet into structured rows."""
    headers = ["Time", "Cur.", "Imp.", "Event", "Actual", "Forecast", "Previous"]
    rows = []
    soup = BeautifulSoup(html_snippet, "html.parser")
    for tr in soup.find_all("tr"):
        the_day_cell = tr.find("td", class_="theDay")
        if the_day_cell:
            date_value = the_day_cell.get_text(strip=True)
            rows.append([date_value] + [""] * (len(headers) - 1))
            continue

        if not tr.get("id", "").startswith("eventRowId_"):
            continue

        columns = tr.find_all("td")
        if not columns:
            continue

        time_value = columns[0].get_text(" ", strip=True) if len(columns) > 0 else ""

        currency_value = (
            columns[1].get_text(" ", strip=True) if len(columns) > 1 else ""
        )

        impact_value = ""
        if len(columns) > 2:
            impact_td = columns[2]
            full_stars = len(impact_td.select("i.grayFullBullishIcon"))
            if full_stars:
                impact_value = {1: "Low", 2: "Medium", 3: "High"}.get(full_stars, "")
            else:
                impact_value = impact_td.get_text(" ", strip=True)

        event_value = columns[3].get_text(" ", strip=True) if len(columns) > 3 else ""
        actual_value = columns[4].get_text(" ", strip=True) if len(columns) > 4 else ""
        forecast_value = (
            columns[5].get_text(" ", strip=True) if len(columns) > 5 else ""
        )
        previous_value = (
            columns[6].get_text(" ", strip=True) if len(columns) > 6 else ""
        )

        row = [
            time_value,
            currency_value,
            impact_value,
            event_value,
            actual_value,
            forecast_value,
            previous_value,
        ]

        while len(row) < len(headers):
            row.append("")

        rows.append(row[: len(headers)])

    return headers, rows


def _post_calendar_with_retries(
    session: requests.Session,
    payload: dict,
    chunk_start: datetime,
    chunk_end: datetime,
):
    cooldowns_used = 0
    session_resets_used = 0
    _reset_raw = (os.getenv("CALENDAR_HTTP_MAX_SESSION_RESETS") or "").strip()
    try:
        max_session_resets = int(_reset_raw or "1")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_HTTP_MAX_SESSION_RESETS={_reset_raw!r}; "
            "falling back to 1."
        )
        max_session_resets = 1

    while True:
        last_status = None
        for attempt in range(HTTP_MAX_ATTEMPTS):
            _sleep_min_interval()
            _sleep_request_jitter()
            _record_request()
            _log_http_stats(
                f"POST {chunk_start:%Y-%m-%d}..{chunk_end:%Y-%m-%d} offset={payload.get('limit_from')} attempt={attempt + 1}",
            )
            response = session.post(
                ECON_CALENDAR_ENDPOINT,
                headers=ECON_CALENDAR_HEADERS,
                data=payload,
                timeout=60,
            )
            if response.status_code == 200:
                _log_http_stats(
                    f"OK {chunk_start:%Y-%m-%d}..{chunk_end:%Y-%m-%d} offset={payload.get('limit_from')}",
                    status=200,
                )
                return response

            last_status = response.status_code
            _log_http_stats(
                f"ERR {chunk_start:%Y-%m-%d}..{chunk_end:%Y-%m-%d} offset={payload.get('limit_from')}",
                status=last_status,
            )
            wait_time = 2**attempt
            print(
                f"[WARNING] Calendar request failed with status {response.status_code}. "
                f"Retrying in {wait_time} seconds..."
            )
            time.sleep(wait_time)

        if last_status == 429 and cooldowns_used < HTTP_RATE_LIMIT_MAX_COOLDOWNS:
            cooldowns_used += 1
            print(
                f"[WARNING] Calendar rate limit persisted (429) for "
                f"{chunk_start:%Y-%m-%d} to {chunk_end:%Y-%m-%d}. "
                f"Cooling down for {HTTP_RATE_LIMIT_COOLDOWN_SECONDS} seconds "
                f"({cooldowns_used}/{HTTP_RATE_LIMIT_MAX_COOLDOWNS})..."
            )
            time.sleep(HTTP_RATE_LIMIT_COOLDOWN_SECONDS)
            continue

        if last_status == 429 and session_resets_used < max_session_resets:
            session_resets_used += 1
            print(
                f"[WARNING] Calendar rate limit persisted (429) for "
                f"{chunk_start:%Y-%m-%d} to {chunk_end:%Y-%m-%d}. "
                "Resetting HTTP session and retrying "
                f"({session_resets_used}/{max_session_resets})..."
            )
            try:
                session.close()
            except Exception:
                pass
            session = requests.Session()
            continue

        raise RuntimeError(
            f"Failed to fetch data for {chunk_start:%Y-%m-%d} to {chunk_end:%Y-%m-%d}"
        )


def fetch_calendar_range(start_date: datetime, end_date: datetime):
    """Fetch calendar data from the calendar provider without Selenium."""
    session = requests.Session()
    headers = ["Time", "Cur.", "Imp.", "Event", "Actual", "Forecast", "Previous"]
    all_rows = []

    _chunk_days_raw = (os.getenv("CALENDAR_RANGE_CHUNK_DAYS") or "").strip()
    try:
        # Smaller chunks reduce pagination depth and lower the risk of partial windows
        # under rate limits. Override via CALENDAR_RANGE_CHUNK_DAYS when needed.
        chunk_days = int(_chunk_days_raw or "4")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_RANGE_CHUNK_DAYS={_chunk_days_raw!r}; "
            "falling back to 4."
        )
        chunk_days = 4

    for chunk_start, chunk_end in chunk_date_range(start_date, end_date, chunk_days):
        try:
            offset = 0
            total_rows_for_chunk = 0
            last_time_scope = None

            while True:
                payload = {
                    "importance[]": list(DEFAULT_IMPORTANCE),
                    "timeZone": DEFAULT_TIMEZONE_ID,
                    "timeFilter": "timeRemain",
                    "currentTab": "custom",
                    "dateFrom": chunk_start.strftime("%Y-%m-%d"),
                    "dateTo": chunk_end.strftime("%Y-%m-%d"),
                    "submitFilters": 1,
                    "limit_from": offset,
                }

                if last_time_scope is not None:
                    payload["last_time_scope"] = last_time_scope

                response = _post_calendar_with_retries(
                    session,
                    payload,
                    chunk_start=chunk_start,
                    chunk_end=chunk_end,
                )

                payload_json = response.json()
                payload_headers, chunk_rows = parse_calendar_html(
                    payload_json.get("data", "")
                )

                if not chunk_rows:
                    if offset == 0:
                        print(
                            f"[INFO] No rows returned for {chunk_start:%Y-%m-%d} to "
                            f"{chunk_end:%Y-%m-%d}."
                        )
                    break

                all_rows.extend(chunk_rows)
                headers = payload_headers

                rows_in_response = int(payload_json.get("rows_num", 0) or 0)
                total_rows_for_chunk += rows_in_response
                last_time_scope = payload_json.get("last_time_scope")
                print(
                    f"[INFO] Retrieved {len(chunk_rows)} rows for "
                    f"{chunk_start:%Y-%m-%d} to {chunk_end:%Y-%m-%d} (offset={offset})."
                )

                bind_scroll = bool(payload_json.get("bind_scroll_handler", False))
                if rows_in_response < ECON_CALENDAR_PAGE_SIZE:
                    if HTTP_STATS_ENABLED:
                        print(
                            f"[INFO] Paging stop: rows_num={rows_in_response} < {ECON_CALENDAR_PAGE_SIZE}"
                        )
                    break
                if not bind_scroll:
                    if HTTP_STATS_ENABLED:
                        print("[INFO] Paging stop: bind_scroll_handler=false")
                    break
                if last_time_scope is None:
                    if HTTP_STATS_ENABLED:
                        print("[INFO] Paging stop: last_time_scope missing")
                    break

                offset += rows_in_response
                time.sleep(
                    random.uniform(PAGE_DELAY_MIN_SECONDS, PAGE_DELAY_MAX_SECONDS)
                )

            if total_rows_for_chunk == 0:
                print(
                    f"[INFO] No data accumulated for {chunk_start:%Y-%m-%d} to "
                    f"{chunk_end:%Y-%m-%d}."
                )

            _sleep_day_delay()
        except Exception as exc:
            raise CalendarFetchError(
                str(exc),
                failed_start=chunk_start,
                failed_end=chunk_end,
                partial_headers=headers,
                partial_rows=all_rows,
            ) from exc

    return headers, all_rows


def save_data(
    headers,
    data,
    file_name="usd_calendar_month.xlsx",
    source_url="",
    *,
    prune_existing_in_range: bool = False,
):
    """Persist calendar data per year without maintaining a master workbook."""
    if not data:
        print("[INFO] No data received. Skipping save operation.")
        return

    print("\n--- Saving Data ---")

    CALENDAR_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    YEARLY_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    headers = ["Date", "Day"] + headers
    processed_data = []
    current_date = None

    for row in data:
        if (
            len(row) > 0
            and row[0].strip()
            and all(col.strip() == "" for col in row[1:])
        ):
            try:
                date_candidate = row[0].strip()
                current_date = datetime.strptime(date_candidate, "%A, %B %d, %Y")
                current_day = current_date.strftime("%A")
            except ValueError:
                current_date = None
                current_day = None
        else:
            if current_date:
                processed_data.append(
                    [current_date.strftime("%Y-%m-%d"), current_day] + row
                )
            else:
                processed_data.append(["", ""] + row)

    df = pd.DataFrame(processed_data, columns=headers)

    while df.columns[0].strip() == "" and df.iloc[:, 0].replace("", None).isna().all():
        df.drop(df.columns[0], axis=1, inplace=True)

    while (
        df.columns[-1].strip() == "" and df.iloc[:, -1].replace("", None).isna().all()
    ):
        df.drop(df.columns[-1], axis=1, inplace=True)

    df.replace(["nan", "NaN", "None"], pd.NA, inplace=True)
    df["Date_dt"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date_dt"]).copy()

    if df.empty:
        print("[WARNING] No valid dated rows found after parsing. Skipping save.")
        return

    df["Day"] = df["Date_dt"].dt.strftime("%A")
    df["Time"] = df["Time"].fillna("").astype(str).str.strip()

    years_to_process = sorted(df["Date_dt"].dt.year.unique())

    for year in years_to_process:
        year_mask = df["Date_dt"].dt.year == year
        year_df = df.loc[year_mask].drop(columns=["Date_dt"]).copy()
        year_dir = os.path.join(YEARLY_OUTPUT_DIR, str(year))
        os.makedirs(year_dir, exist_ok=True)

        excel_path = Path(year_dir) / f"{year}_calendar.xlsx"

        if excel_path.exists():
            existing_df = pd.read_excel(excel_path, sheet_name="Data")
        else:
            existing_df = pd.DataFrame(columns=year_df.columns)

        for col in year_df.columns:
            if col not in existing_df.columns:
                existing_df[col] = pd.NA

        existing_df = existing_df[year_df.columns]

        existing_df_compare = existing_df.copy()
        existing_df_merge = existing_df
        if prune_existing_in_range and not year_df.empty and "Date" in year_df.columns:
            year_dates = pd.to_datetime(year_df["Date"], errors="coerce")
            if year_dates.notna().any():
                start_date = year_dates.min().strftime("%Y-%m-%d")
                end_date = year_dates.max().strftime("%Y-%m-%d")
                # Guard against upstream/API anomalies: prune per-day only when
                # the newly fetched window looks reasonably complete.
                safe_prune_days, skipped_prune_days = (
                    calendar_pruning.compute_safe_prune_days(
                        existing_df_compare,
                        year_df,
                        start_date,
                        end_date,
                        guard_ratio=PRUNE_GUARD_RATIO,
                        guard_min_new_nonholiday=PRUNE_GUARD_MIN_NEW_NONHOLIDAY,
                    )
                )
                for day in sorted(skipped_prune_days):
                    print(
                        f"[WARNING] Skipping prune for {day}: {skipped_prune_days[day]}"
                    )

                if safe_prune_days:
                    date_str = existing_df_merge.get("Date")
                    if date_str is not None:
                        before_rows = len(existing_df_merge)
                        mask = date_str.fillna("").astype(str).isin(safe_prune_days)
                        existing_df_merge = existing_df_merge.loc[~mask].copy()
                        pruned_rows = before_rows - len(existing_df_merge)
                        print(
                            f"[INFO] Pruned {pruned_rows} existing rows inside "
                            f"{start_date}..{end_date} (safe_days={len(safe_prune_days)})."
                        )

        combined_df = processing.merge_calendar_frames(existing_df_merge, year_df)

        combined_sorted = processing.sort_calendar_dataframe(
            combined_df.copy()
        ).reset_index(drop=True)

        existing_sorted = (
            processing.sort_calendar_dataframe(existing_df_compare.copy())
            .reindex(columns=combined_sorted.columns)
            .reset_index(drop=True)
        )

        existing_norm = processing.normalize_calendar_frame_for_compare(existing_sorted)
        combined_norm = processing.normalize_calendar_frame_for_compare(combined_sorted)

        if existing_norm.equals(combined_norm):
            print(f"[INFO] Year {year} unchanged; skipping write.")
            continue

        processing.write_calendar_outputs(combined_sorted, excel_path)

        print(f"[SUCCESS] Year {year} exports written to {excel_path.parent}")


def resolve_prune_existing_in_range(args) -> bool:
    """Resolve pruning toggle from args/env with a safe default."""
    if getattr(args, "prune_existing_in_range", None) is not None:
        return bool(args.prune_existing_in_range)
    env_raw = (os.getenv("CALENDAR_PRUNE_EXISTING_IN_RANGE") or "").strip().lower()
    if env_raw in {"0", "false", "no", "off"}:
        return False
    if env_raw in {"1", "true", "yes", "on"}:
        return True
    # Default: enabled, but protected by per-day prune guard.
    return True


def main():
    """Main script steps using the HTTP calendar endpoint."""
    args = parse_args()

    start_date, end_date = resolve_date_range(args)

    print(
        f"[INFO] Fetching economic calendar data from {start_date:%Y-%m-%d} "
        f"to {end_date:%Y-%m-%d} via HTTP API."
    )

    try:
        headers, data = fetch_calendar_range(start_date, end_date)
    except CalendarFetchError as exc:
        print(f"[ERROR] {exc}")
        if exc.partial_rows:
            print(
                f"[WARNING] Saving {len(exc.partial_rows)} rows fetched before "
                f"failure at {exc.failed_start:%Y-%m-%d}..{exc.failed_end:%Y-%m-%d}."
            )
            save_data(
                exc.partial_headers,
                exc.partial_rows,
                file_name="usd_calendar_month.xlsx",
                source_url=CALENDAR_REFERER,
                prune_existing_in_range=False,
            )
        sys.exit(1)
    except Exception as exc:
        print(f"[ERROR] {exc}")
        sys.exit(1)

    if not data:
        print("[WARNING] No data fetched. Skipping save operation.")
        return

    try:
        save_data(
            headers,
            data,
            file_name="usd_calendar_month.xlsx",
            source_url=CALENDAR_REFERER,
            prune_existing_in_range=resolve_prune_existing_in_range(args),
        )
    except Exception as exc:
        print(f"[ERROR] Failed to save calendar data: {exc}")
        sys.exit(1)


if __name__ == "__main__":
    main()
