import os
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

_URL_TOKEN_RE = re.compile(r"(?:https?://|www\\.)\\S+", re.IGNORECASE)
_DOMAIN_TOKEN_RE = re.compile(
    r"\\b(?:[A-Za-z0-9-]{2,63}\\.)+[A-Za-z]{2,24}\\b", re.IGNORECASE
)

COLUMN_WIDTH_OVERRIDES = {
    "Date": 9.71,
    "Day": 10.71,
    "Time": 4.86,
    "Cur.": 4.57,
    "Imp.": 7.71,
    "Event": 50.0,
    "Actual": 8.57,
    "Forecast": 8.57,
    "Previous": 8.57,
}

KEY_COLUMNS = ["Date", "Time", "Cur.", "Event"]
MISSING_VALUE_TOKENS = {"tba", "tentative", "n/a", "na"}
VALUE_COLUMNS = ["Actual", "Forecast", "Previous"]
_MONTH_SUFFIX_RE = re.compile(
    r"\s*\(\s*(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*\)\s*$",
    re.IGNORECASE,
)


def _has_month_suffix(event_name: str) -> bool:
    return bool(_MONTH_SUFFIX_RE.search(event_name.strip()))


def _strip_month_suffix(event_name: str) -> str:
    # Collapse provider whitespace and remove trailing "(Nov)" style tags.
    cleaned = re.sub(r"\s{2,}", " ", str(event_name or "").strip())
    return _MONTH_SUFFIX_RE.sub("", cleaned).strip()


def _event_datetime_source_tz(date_str: str, time_str: str) -> datetime | None:
    if not date_str:
        return None
    try:
        day = datetime.strptime(date_str.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None

    time_text = (time_str or "").strip()
    if ":" in time_text:
        try:
            t = datetime.strptime(time_text, "%H:%M").time()
        except ValueError:
            t = datetime.min.time()
    else:
        # Treat all-day / missing time as end-of-day for "not in future" pruning.
        t = datetime.strptime("23:59", "%H:%M").time()

    source_tz = timezone(timedelta(hours=8))
    return datetime.combine(day, t).replace(tzinfo=source_tz)


def _sanitize_text_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    cleaned = _URL_TOKEN_RE.sub("", value)
    cleaned = _DOMAIN_TOKEN_RE.sub("", cleaned)
    cleaned = re.sub(r"\\s{2,}", " ", cleaned).strip()
    return cleaned


def _parse_time_minutes(value: object) -> float | None:
    if not isinstance(value, str):
        return None
    text = value.strip()
    try:
        parsed = datetime.strptime(text, "%H:%M")
    except ValueError:
        return None
    return float(parsed.hour * 60 + parsed.minute)


def _format_time_minutes(minutes: float) -> str:
    total = int(round(minutes))
    total = max(0, min(total, 23 * 60 + 59))
    hour = total // 60
    minute = total % 60
    return f"{hour:02d}:{minute:02d}"


def _is_missing_token(value: object, *, missing_tokens: set[str]) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    text = str(value).strip()
    if not text:
        return True
    return text.lower() in missing_tokens


def _normalize_missing_text_values(
    df: pd.DataFrame, *, missing_tokens: set[str]
) -> pd.DataFrame:
    """Normalize missing tokens for non-key text columns.

    We keep KEY_COLUMNS as normalized strings (they are used for identity), but
    treat empty strings / provider placeholders as missing for the remaining
    fields so JSON exports are stable (missing -> null) across fetch runs.
    """
    working = df.copy()
    for col_name in working.columns:
        if col_name in KEY_COLUMNS:
            continue
        if working[col_name].dtype != object:
            continue
        text = working[col_name].astype(str).str.strip()
        is_missing = (
            working[col_name].isna()
            | text.eq("")
            | text.str.lower().isin(missing_tokens)
        )
        working.loc[is_missing, col_name] = pd.NA
        working.loc[~is_missing, col_name] = text.loc[~is_missing]
    return working


def _format_number_for_compare(value: float) -> str:
    # Avoid "113.0" vs "113" churn when reading back from Excel.
    if float(value).is_integer():
        return str(int(value))
    return format(float(value), "g")


def _normalize_value_for_compare(value: object, *, missing_tokens: set[str]) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return _format_number_for_compare(float(value))
    text = str(value).strip()
    if not text:
        return ""
    if text.lower() in missing_tokens:
        return ""
    return text


def normalize_calendar_frame_for_compare(df: pd.DataFrame) -> pd.DataFrame:
    """Return a fully-string DataFrame suitable for stable equality checks."""
    working = df.copy()
    working.replace(["nan", "NaN", "None"], pd.NA, inplace=True)
    for col_name in working.columns:
        if col_name in KEY_COLUMNS:
            working[col_name] = working[col_name].fillna("").astype(str).str.strip()
            continue
        series = working[col_name]
        working[col_name] = [
            _normalize_value_for_compare(v, missing_tokens=MISSING_VALUE_TOKENS)
            for v in series.tolist()
        ]
    return working


def prune_calendar_frame_by_date_range(
    df: pd.DataFrame, *, start_date: str, end_date: str
) -> pd.DataFrame:
    """Drop rows whose Date falls within [start_date, end_date] (inclusive)."""
    if df.empty or "Date" not in df.columns:
        return df.copy()
    start_dt = pd.to_datetime(start_date, errors="coerce")
    end_dt = pd.to_datetime(end_date, errors="coerce")
    if pd.isna(start_dt) or pd.isna(end_dt):
        return df.copy()
    date_dt = pd.to_datetime(df["Date"], errors="coerce")
    in_range = date_dt.notna() & (date_dt >= start_dt) & (date_dt <= end_dt)
    return df.loc[~in_range].copy()


def _normalize_value_token(value: object, *, missing_tokens: set[str]) -> str:
    if _is_missing_token(value, missing_tokens=missing_tokens):
        return ""
    return str(value).strip()


def _choose_canonical_time(times: list[float]) -> float | None:
    if not times:
        return None

    def score(t: float) -> tuple[int, int, float]:
        minute = int(round(t)) % 60
        is_hour_or_half = 1 if minute in {0, 30} else 0
        is_quarter = 1 if minute in {0, 15, 30, 45} else 0
        return (is_hour_or_half, is_quarter, t)

    return max(times, key=score)


def _snap_time_to_canonical_minutes(minutes: float, *, threshold_minutes: int) -> float:
    if threshold_minutes <= 0:
        return minutes
    if pd.isna(minutes):
        return minutes
    rounded = int(round(minutes))
    if rounded < 0 or rounded > 23 * 60 + 59:
        return minutes

    canonical = list(range(0, 24 * 60, 5))
    nearest = min(canonical, key=lambda m: abs(m - rounded))
    if abs(nearest - rounded) <= threshold_minutes:
        return float(nearest)
    return minutes


def sort_calendar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Sort ``df`` by date and the semantic meaning of the ``Time`` column."""
    if df.empty or "Date" not in df.columns or "Time" not in df.columns:
        return df

    working = df.copy()
    working["_sort_date"] = pd.to_datetime(working["Date"], errors="coerce")
    time_str = working["Time"].fillna("").astype(str).str.strip()

    working["_is_all_day"] = time_str.str.lower() == "all day"
    working["_time_parsed"] = pd.to_datetime(time_str, format="%H:%M", errors="coerce")

    working["_sort_bucket"] = 2  # textual placeholders (e.g. Tentative/TBA)
    working.loc[working["_is_all_day"], "_sort_bucket"] = 0
    working.loc[working["_time_parsed"].notna(), "_sort_bucket"] = 1
    working.loc[
        time_str.eq("") | time_str.str.lower().isin(["nan", "none"]), "_sort_bucket"
    ] = 3

    tie_breakers = [
        "_sort_date",
        "_sort_bucket",
        "_time_parsed",
        "Cur.",
        "Imp.",
        "Event",
    ]
    sort_columns = [c for c in tie_breakers if c in working.columns]
    ascending_flags = [True] * len(sort_columns)
    working.sort_values(
        by=sort_columns,
        ascending=ascending_flags,
        inplace=True,
        kind="stable",
    )

    working.drop(
        columns=["_sort_date", "_sort_bucket", "_time_parsed", "_is_all_day"],
        inplace=True,
        errors="ignore",
    )
    return working


def _apply_fuzzy_time_dedup(
    df: pd.DataFrame,
    *,
    group_columns: list[str],
    threshold_minutes: int,
) -> pd.DataFrame:
    if df.empty or threshold_minutes <= 0:
        return df

    if any(col not in df.columns for col in group_columns):
        return df

    working = df.copy()
    working["_time_minutes"] = working.get("Time", "").map(_parse_time_minutes)

    missing_tokens = {"tba", "tentative", "n/a", "na"}

    kept_rows: list[pd.Series] = []
    for _, group in working.groupby(group_columns, dropna=False, sort=False):
        if len(group) == 1:
            row = group.iloc[0].copy()
            row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(row)
            continue

        timed = group[group["_time_minutes"].notna()].copy()
        untimed = group[group["_time_minutes"].isna()].copy()

        if timed.empty:
            for _, row in group.iterrows():
                row = row.copy()
                row.drop(labels=["_time_minutes"], inplace=True)
                kept_rows.append(row)
            continue

        timed = timed.sort_values(by="_time_minutes", kind="mergesort")
        clusters: list[list[int]] = []
        current: list[int] = []
        last_time: float | None = None
        for idx, row in timed.iterrows():
            t = row["_time_minutes"]
            if not isinstance(t, float) or pd.isna(t):
                continue
            if last_time is None or abs(t - last_time) <= threshold_minutes:
                current.append(int(idx))
            else:
                clusters.append(current)
                current = [int(idx)]
            last_time = t
        if current:
            clusters.append(current)

        for cluster in clusters:
            cluster_df = timed.loc[cluster].copy()
            best_idx = int(
                cluster_df.sort_values(
                    by=["completeness_score"],
                    ascending=False,
                    kind="mergesort",
                ).index[0]
            )
            best_row = cluster_df.loc[best_idx].copy()

            canonical_time = _choose_canonical_time(
                [
                    float(x)
                    for x in cluster_df["_time_minutes"].tolist()
                    if isinstance(x, float) and not pd.isna(x)
                ]
            )
            if canonical_time is not None:
                best_row["Time"] = _format_time_minutes(canonical_time)

            for _, candidate in cluster_df.iterrows():
                if int(candidate.name) == best_idx:
                    continue
                for col_name in cluster_df.columns:
                    if col_name in {"_time_minutes"}:
                        continue
                    if _is_missing_token(
                        best_row.get(col_name), missing_tokens=missing_tokens
                    ) and not _is_missing_token(
                        candidate.get(col_name), missing_tokens=missing_tokens
                    ):
                        best_row[col_name] = candidate.get(col_name)

            best_row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(best_row)

        for _, row in untimed.iterrows():
            row = row.copy()
            row.drop(labels=["_time_minutes"], inplace=True)
            kept_rows.append(row)

    return pd.DataFrame(kept_rows).reset_index(drop=True)


def _apply_update_slot_dedup(
    df: pd.DataFrame,
    *,
    threshold_minutes: int,
) -> pd.DataFrame:
    """Merge near-duplicate "update slot" rows.

    Some events appear twice on the same day with the same Forecast/Previous:
    - Earlier row: Actual missing (UI shows TBA)
    - Later row: Actual filled
    If the times are close enough, keep the row with Actual and merge any
    missing values from the other row, then drop the duplicate.
    """
    if df.empty or threshold_minutes <= 0:
        return df

    required = {"Date", "Cur.", "Event", "Time", "Actual", "Forecast", "Previous"}
    if any(col not in df.columns for col in required):
        return df

    working = df.copy()
    missing_tokens = {"tba", "tentative", "n/a", "na"}

    keep_mask = pd.Series(True, index=working.index)
    # Group by Cur./Event so we can merge across adjacent-day updates
    # (e.g. 23:30 placeholder then 00:30 actual).
    for _, group in working.groupby(["Cur.", "Event"], dropna=False, sort=False):
        if len(group) < 2:
            continue

        timed = group.copy()
        timed["_time_minutes"] = timed["Time"].map(_parse_time_minutes)
        timed["_date_value"] = pd.to_datetime(timed["Date"], errors="coerce").dt.date
        timed = timed[
            timed["_time_minutes"].notna() & timed["_date_value"].notna()
        ].copy()
        if len(timed) < 2:
            continue

        timed["_ts_minutes"] = (
            timed["_date_value"].map(lambda d: d.toordinal()) * (24 * 60)
            + timed["_time_minutes"]
        )
        timed.sort_values(by="_ts_minutes", inplace=True, kind="mergesort")
        indices = timed.index.tolist()

        for i, idx_i in enumerate(indices):
            if not keep_mask.get(idx_i, False):
                continue
            row_i = working.loc[idx_i]
            ts_i = timed.loc[idx_i, "_ts_minutes"]
            if not isinstance(ts_i, float) or pd.isna(ts_i):
                continue

            for idx_j in indices[i + 1 :]:
                if not keep_mask.get(idx_j, False):
                    continue
                ts_j = timed.loc[idx_j, "_ts_minutes"]
                if not isinstance(ts_j, float) or pd.isna(ts_j):
                    continue
                if ts_j - ts_i > threshold_minutes:
                    break

                row_j = working.loc[idx_j]

                prev_i = _normalize_value_token(
                    row_i.get("Previous"), missing_tokens=missing_tokens
                )
                prev_j = _normalize_value_token(
                    row_j.get("Previous"), missing_tokens=missing_tokens
                )
                fc_i = _normalize_value_token(
                    row_i.get("Forecast"), missing_tokens=missing_tokens
                )
                fc_j = _normalize_value_token(
                    row_j.get("Forecast"), missing_tokens=missing_tokens
                )
                # Treat missing forecast/previous as compatible, but if both
                # sides have values they must match.
                if prev_i and prev_j and prev_i != prev_j:
                    continue
                if fc_i and fc_j and fc_i != fc_j:
                    continue

                actual_i_missing = _is_missing_token(
                    row_i.get("Actual"), missing_tokens=missing_tokens
                )
                actual_j_missing = _is_missing_token(
                    row_j.get("Actual"), missing_tokens=missing_tokens
                )
                if actual_i_missing == actual_j_missing:
                    continue

                if actual_i_missing and not actual_j_missing:
                    keep_idx, drop_idx = idx_j, idx_i
                else:
                    keep_idx, drop_idx = idx_i, idx_j

                keep_row = working.loc[keep_idx].copy()
                drop_row = working.loc[drop_idx]
                for col in working.columns:
                    if col in {"_time_minutes"}:
                        continue
                    if _is_missing_token(
                        keep_row.get(col), missing_tokens=missing_tokens
                    ) and not _is_missing_token(
                        drop_row.get(col), missing_tokens=missing_tokens
                    ):
                        keep_row[col] = drop_row.get(col)

                working.loc[keep_idx] = keep_row
                keep_mask[drop_idx] = False

                if drop_idx == idx_i:
                    row_i = working.loc[keep_idx]
                    ts_i = ts_j
                    idx_i = keep_idx

    return working.loc[keep_mask].copy()


def merge_calendar_frames(
    existing_df: pd.DataFrame, new_df: pd.DataFrame
) -> pd.DataFrame:
    existing_tagged = existing_df.copy()
    existing_tagged["_source_rank"] = 0
    new_tagged = new_df.copy()
    new_tagged["_source_rank"] = 1
    working = pd.concat([existing_tagged, new_tagged], ignore_index=True, sort=False)
    working.replace(["nan", "NaN", "None"], pd.NA, inplace=True)

    # Normalize key fields to avoid duplicates differing only by NA vs "".
    # Some provider rows (e.g., holidays) may omit `Cur.`; pandas writes NA as
    # blank in CSV, which can reintroduce exact duplicates across runs.
    for col_name in KEY_COLUMNS:
        if col_name in working.columns:
            working[col_name] = working[col_name].fillna("").astype(str).str.strip()
        else:
            working[col_name] = ""

    for col_name in working.columns:
        if working[col_name].dtype == object:
            working[col_name] = working[col_name].map(_sanitize_text_value)

    # Keep value columns as text to avoid Excel re-read type churn.
    for col_name in VALUE_COLUMNS:
        if col_name not in working.columns:
            continue
        working[col_name] = [
            (
                pd.NA
                if _normalize_value_for_compare(v, missing_tokens=MISSING_VALUE_TOKENS)
                == ""
                else _normalize_value_for_compare(
                    v, missing_tokens=MISSING_VALUE_TOKENS
                )
            )
            for v in working[col_name].tolist()
        ]

    working = _normalize_missing_text_values(
        working, missing_tokens=MISSING_VALUE_TOKENS
    )

    working["Date_dt"] = pd.to_datetime(working.get("Date"), errors="coerce")
    working = working.dropna(subset=["Date_dt"])  # type: ignore[arg-type]

    completeness = pd.Series(0, index=working.index, dtype="int64")
    for col_name in working.columns:
        col = working[col_name]
        present = col.notna()
        if col.dtype == object:
            text = col.astype(str).str.strip()
            present &= text.ne("") & ~text.str.lower().isin(MISSING_VALUE_TOKENS)
        completeness += present.astype("int64")

    working["completeness_score"] = completeness
    # Stable ordering: prefer more complete rows, then prefer new fetch data.
    working = working.sort_values(
        by=["completeness_score", "_source_rank"],
        ascending=[False, False],
        kind="mergesort",
    )
    working = working.drop_duplicates(subset=KEY_COLUMNS, keep="first")

    _fuzzy_raw = (os.getenv("CALENDAR_EVENT_TIME_FUZZY_DEDUP_MINUTES") or "").strip()
    try:
        fuzzy_minutes = int(_fuzzy_raw or "2")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_EVENT_TIME_FUZZY_DEDUP_MINUTES={_fuzzy_raw!r}; "
            "falling back to 2."
        )
        fuzzy_minutes = 2

    working = _apply_fuzzy_time_dedup(
        working,
        group_columns=["Date", "Cur.", "Event"],
        threshold_minutes=fuzzy_minutes,
    )

    _update_raw = (os.getenv("CALENDAR_UPDATE_SLOT_DEDUP_MINUTES") or "").strip()
    try:
        update_minutes = int(_update_raw or "60")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_UPDATE_SLOT_DEDUP_MINUTES={_update_raw!r}; "
            "falling back to 60."
        )
        update_minutes = 60

    working = _apply_update_slot_dedup(working, threshold_minutes=update_minutes)
    working = _drop_stale_month_placeholder_rows(working)
    working.drop(columns=["completeness_score", "_source_rank"], inplace=True)

    _snap_raw = (os.getenv("CALENDAR_TIME_SNAP_THRESHOLD_MINUTES") or "").strip()
    try:
        snap_threshold = int(_snap_raw or "2")
    except ValueError:
        print(
            f"[WARNING] Invalid CALENDAR_TIME_SNAP_THRESHOLD_MINUTES={_snap_raw!r}; "
            "falling back to 2."
        )
        snap_threshold = 2

    if snap_threshold > 0 and "Time" in working.columns:
        minutes = working["Time"].map(_parse_time_minutes)
        working["Time"] = [
            (
                _format_time_minutes(
                    _snap_time_to_canonical_minutes(m, threshold_minutes=snap_threshold)
                )
                if isinstance(m, float) and not pd.isna(m)
                else t
            )
            for t, m in zip(working["Time"].tolist(), minutes.tolist())
        ]

    working["Date"] = working["Date_dt"].dt.strftime("%Y-%m-%d")
    working["Day"] = working["Date_dt"].dt.strftime("%A")
    working.drop(columns=["Date_dt"], inplace=True)

    working = sort_calendar_dataframe(working)
    return working.reset_index(drop=True)


def _drop_stale_month_placeholder_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Drop stale placeholder duplicates that can show as TBA in the UI.

    We only drop rows when:
    - Same Cur., Imp., Previous
    - Same month (YYYY-MM in Date)
    - One row has Actual, another row is missing Actual
    - And either:
      - Event names are identical, OR
      - Event names match after removing a trailing "(Nov)" month tag and exactly one row has it
    - Both rows are not in the future relative to UTC+8 (calendar source timezone)

    This avoids merging genuine month-to-month rows like \"(Aug)\" vs \"(Sep)\".
    """
    required = {"Date", "Time", "Cur.", "Imp.", "Event", "Actual", "Previous"}
    if df.empty or not required.issubset(set(df.columns)):
        return df.copy()

    working = df.copy()

    # Normalize key-ish columns used for matching.
    for col in ["Cur.", "Imp.", "Previous", "Event"]:
        working[col] = working[col].fillna("").astype(str).str.strip()
    working["Actual"] = working["Actual"].fillna("").astype(str).str.strip()

    working["_ym"] = working["Date"].astype(str).str.slice(0, 7)
    working["_event_exact"] = working["Event"].map(lambda v: re.sub(r"\s{2,}", " ", v).strip())
    working["_event_base"] = working["_event_exact"].map(_strip_month_suffix)
    working["_has_month"] = working["_event_exact"].map(_has_month_suffix)

    now_source = datetime.now(timezone(timedelta(hours=8)))
    dts = [
        _event_datetime_source_tz(d, t)
        for d, t in zip(working["Date"].astype(str).tolist(), working["Time"].astype(str).tolist())
    ]
    working["_dt_source"] = dts
    working["_not_future"] = working["_dt_source"].map(
        lambda dt: bool(dt) and dt <= now_source
    )

    def actual_missing(series: pd.Series) -> pd.Series:
        text = series.astype(str).str.strip().str.lower()
        return text.eq("") | text.isin(MISSING_VALUE_TOKENS) | text.isin({"--", "-", "\u2014", "null"})

    working["_actual_missing"] = actual_missing(working["Actual"])
    working["_actual_present"] = ~working["_actual_missing"]

    drop_mask = pd.Series(False, index=working.index)

    group_cols_exact = ["_ym", "Cur.", "Imp.", "Previous", "_event_exact"]
    exact_groups = working.groupby(group_cols_exact, dropna=False, sort=False)
    for _, idxs in exact_groups.groups.items():
        group = working.loc[list(idxs)]
        if not group["_not_future"].any():
            continue
        present = group[group["_actual_present"] & group["_not_future"]]
        if present.empty:
            continue
        missing = group[group["_actual_missing"] & group["_not_future"]]
        if missing.empty:
            continue
        drop_mask.loc[missing.index] = True

    group_cols_base = ["_ym", "Cur.", "Imp.", "Previous", "_event_base"]
    base_groups = working.groupby(group_cols_base, dropna=False, sort=False)
    for _, idxs in base_groups.groups.items():
        group = working.loc[list(idxs)]
        if group["_has_month"].nunique(dropna=False) < 2:
            continue
        if not group["_not_future"].any():
            continue

        month_actual = group[
            group["_has_month"] & group["_actual_present"] & group["_not_future"]
        ]
        nomonth_actual = group[
            (~group["_has_month"]) & group["_actual_present"] & group["_not_future"]
        ]

        month_missing = group[
            group["_has_month"] & group["_actual_missing"] & group["_not_future"]
        ]
        nomonth_missing = group[
            (~group["_has_month"]) & group["_actual_missing"] & group["_not_future"]
        ]

        # Drop only the side that has missing Actual when the opposite side has a value.
        if not month_actual.empty and not nomonth_missing.empty:
            drop_mask.loc[nomonth_missing.index] = True
        if not nomonth_actual.empty and not month_missing.empty:
            drop_mask.loc[month_missing.index] = True

    cleaned = working.loc[~drop_mask].copy()
    cleaned.drop(
        columns=[
            "_ym",
            "_event_exact",
            "_event_base",
            "_has_month",
            "_dt_source",
            "_not_future",
            "_actual_missing",
            "_actual_present",
        ],
        inplace=True,
        errors="ignore",
    )
    return cleaned


def write_calendar_outputs(df: pd.DataFrame, excel_path: Path) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Data")
        worksheet = writer.sheets["Data"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            column_letter = get_column_letter(col_idx)
            max_text_length = df[col_name].astype(str).map(len).max() or 0
            override_width = COLUMN_WIDTH_OVERRIDES.get(col_name)
            col_width = (
                override_width
                if override_width is not None
                else max(max_text_length + 2, 15)
            )

            cell_alignment = (
                Alignment(horizontal="left")
                if col_name == "Event"
                else Alignment(horizontal="center")
            )

            worksheet.column_dimensions[column_letter].width = col_width
            for cell in worksheet[column_letter]:
                cell.alignment = cell_alignment

        worksheet.freeze_panes = "A2"

        highlight_fill = PatternFill(
            start_color="FFD700", end_color="FFD700", fill_type="solid"
        )
        previous_date = None
        for row_idx, row in enumerate(df.itertuples(), start=2):
            current_row_date = row.Date
            if current_row_date != previous_date:
                for col_idx in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_idx, column=col_idx).fill = highlight_fill
                previous_date = current_row_date

    df.to_csv(excel_path.with_suffix(".csv"), index=False)
    df.to_json(excel_path.with_suffix(".json"), orient="records", indent=4)
