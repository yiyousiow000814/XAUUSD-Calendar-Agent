from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

_URL_TOKEN_RE = re.compile(r"(?:https?://|www\\.)\\S+", re.IGNORECASE)
_DOMAIN_TOKEN_RE = re.compile(
    r"\\b(?:[A-Za-z0-9-]{2,63}\\.)+[A-Za-z]{2,24}\\b", re.IGNORECASE
)


def sanitize_text(value: str) -> str:
    cleaned = _URL_TOKEN_RE.sub("", value)
    cleaned = _DOMAIN_TOKEN_RE.sub("", cleaned)
    cleaned = re.sub(r"\\s{2,}", " ", cleaned).strip()
    return cleaned


def sanitize_csv(path: Path, dry_run: bool) -> bool:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        fieldnames = reader.fieldnames or []
        rows = list(reader)

    changed = False
    for row in rows:
        for key, raw_value in row.items():
            if not isinstance(raw_value, str) or not raw_value.strip():
                continue
            sanitized = sanitize_text(raw_value)
            if sanitized != raw_value:
                row[key] = sanitized
                changed = True

    if changed and not dry_run:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    return changed


def sanitize_json(path: Path, dry_run: bool) -> bool:
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(payload, list):
        return False

    changed = False
    for item in payload:
        if not isinstance(item, dict):
            continue
        for key, raw_value in item.items():
            if not isinstance(raw_value, str) or not raw_value.strip():
                continue
            sanitized = sanitize_text(raw_value)
            if sanitized != raw_value:
                item[key] = sanitized
                changed = True

    if changed and not dry_run:
        path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=4, separators=(",", ":"))
            + "\n",
            encoding="utf-8",
        )

    return changed


def sanitize_xlsx(path: Path, dry_run: bool) -> bool:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active

        header_row = None
        for row_idx in range(1, 6):
            values = [sheet.cell(row=row_idx, column=col).value for col in range(1, 25)]
            if any(isinstance(v, str) and v.strip() == "Event" for v in values):
                header_row = row_idx
                break
        if header_row is None:
            return False

        event_col = None
        for col_idx in range(1, 25):
            if sheet.cell(row=header_row, column=col_idx).value == "Event":
                event_col = col_idx
                break
        if event_col is None:
            return False

        changed = False
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            raw_value = sheet.cell(row=row_idx, column=event_col).value
            if not isinstance(raw_value, str) or not raw_value.strip():
                continue
            sanitized = sanitize_text(raw_value)
            if sanitized != raw_value:
                sheet.cell(row=row_idx, column=event_col).value = sanitized
                changed = True

        if changed and not dry_run:
            workbook.save(path)
        return changed
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sanitize economic calendar exports in-place (strip domain-like prefixes from Event names)."
    )
    parser.add_argument(
        "--root",
        type=Path,
        default=Path("data") / "Economic_Calendar",
        help="Root directory containing yearly calendar exports.",
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="Report changes without writing."
    )
    args = parser.parse_args()

    root: Path = args.root
    if not root.exists():
        raise SystemExit(f"Calendar directory not found: {root}")

    changed_paths: list[Path] = []
    for path in sorted(root.rglob("*_calendar.*")):
        if path.suffix.lower() == ".csv":
            changed = sanitize_csv(path, args.dry_run)
        elif path.suffix.lower() == ".json":
            changed = sanitize_json(path, args.dry_run)
        elif path.suffix.lower() == ".xlsx":
            changed = sanitize_xlsx(path, args.dry_run)
        else:
            continue
        if changed:
            changed_paths.append(path)

    for path in changed_paths:
        print(path.as_posix())
    print(f"Changed files: {len(changed_paths)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
